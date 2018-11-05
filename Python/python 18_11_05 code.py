# -*- coding: utf-8 -*-
"""
Python 18_11_05

"""

# 맞교환

a = 20
b = 30

a = b
a

k = 'Korea'
j = 'Japan'

k, j = j, k
k
j

# 원소 넣기 / packing, unpacking

d = (1, 2, 3, 4)
t = d
t
t1, t2, t3, t4 = t
print(t1, t2, t3, t4)

# 데코레이터, @staticmethod, @classmethod
class Date(object):
    def __init__(self, year, month, day):
        self.year = year
        self.month = month
        self.day = day
        print(" Date__init__ ")
        
    def __str__(self):
        return "{}년 {}월 {}일 생산".format(self.year, self.month, self.day)
    
class Ktx:
    def __init__(self, station, fee):
        self.station = station
        self.fee = fee
        print("인수 2개 __init__")
        
    def __del__(self):
        print(" __del__ ")
        
    def __str__(self):
        return " {} *** {}원 ".format(self.station, self.fee)

k1 = Ktx("부산역", 49000)
k2 = Ktx("대전역", 24000)

print(k1)
print(k2)

k3 = Ktx() ## error. station, fee argument를 안 받았기 때문.

class Ktx:
    def __init__(self, station, fee):
        self.station = station
        self.fee = fee
        print("인수 2개 __init__")
    # 다음 부분 추가    
    def __init__(self):
        self.station = "이름 없음"
        self.fee = 0
        print("인수 없음 __init__")
        
    def __del__(self):
        print(" __del__ ")
        
    def __str__(self):
        return " {} *** {}원 ".format(self.station, self.fee)
    
k3 = Ktx()

print(k3)

# 합쳐서 다음과 같이 표현 가능
class Ktx:
    def __init__(self, station = "이름 없음", fee = 0): ## Default 설정
        self.station = station
        self.fee = fee
        print("인수 2개 __init__")
        
    def __del__(self):
        print(" __del__ ")
        
    def __str__(self):
        return " {} *** {}원 ".format(self.station, self.fee)
    
k4 = Ktx() # 인자를 하나도 안 준 경우 default
print(k4)


k5 = Ktx("울산역") # 하나만 준 경우 할당된 인자 외에 default
print(k5)

k5.station
a = k5.station
a

# get, set 추가
class Ktx:
    def __init__(self, station = "이름 없음", fee = 0): ## Default 설정
        self.station = station
        self.fee = fee
        print("인수 2개 __init__")
        
    def __del__(self):
        print(" __del__ ")
        
    def __str__(self):
        return " {} *** {}원 ".format(self.station, self.fee)
    
    def get_station(self):
        return self.station
    
    def get_fee(self):
        return self.fee
    
    def set_station(self, station):
        self.station = station
    
    def set_fee(self, fee):
        self.fee = fee


k1 = Ktx("부산역", 49000)
k2 = Ktx("대전역", 24000)
k2.set_station("목포역")
k2.set_fee(47000)
print(k1, k2)

# cnt 추가(class 변수)
## classmethod(객체 하나하나에 종속되는 것이 아닌 class 전체에서 사용하는 method)
class Ktx:
    cnt = 0  ## class 변수
    
    @staticmethod
    def get_cnt_st():
        print("현재 객체수는 : ", Ktx.cnt)
    
    @classmethod
    def get_cnt(cls):
        print("현재 객체 수는 : ", Ktx.cnt)
    
    def __init__(self, station = "이름 없음", fee = 0): ## Default 설정
        self.station = station
        self.fee = fee
        Ktx.cnt += 1
        print("인수 2개 __init__")
        
    def __del__(self):
        Ktx.cnt -= 1
        print(" __del__ ")
        
    def __str__(self):
        return " {} *** {}원 ".format(self.station, self.fee)
    
    def get_station(self):
        return self.station
    
    def get_fee(self):
        return self.fee
    
    def set_station(self, station):
        self.station = station
    
    def set_fee(self, fee):
        self.fee = fee

k1 = Ktx("부산역", 49000)
k2 = Ktx("대전역", 24000)
k3 = Ktx()
k4 = Ktx("울산역")

Ktx.cnt
k1.cnt ## 객체를 통해서도 class 변수에 접근 가능( 왜?? )
Ktx.get_cnt()

del Ktx, k1, k2, k3, k4


# comprehension method

import sys

a = 100
b = 200
aa = [2, 4, 6]
bb = ["kbs", "mbc", "sbs"]

sys.getsizeof(a)
sys.getsizeof(b)
sys.getsizeof(aa)
sys.getsizeof(bb)

aa = []
len(aa)
for i in range(20):
    aa.append(i + 5)

aa
sys.getsizeof(aa)

jj = [j + 5 for j in range(10)]
jj
sys.getsizeof(jj)

kk = (i for i in range(20))
kk

kk.__next__()
sys.getsizeof(kk)
next(kk)

jj = [j for j in range(100)]
jj
j3 = iter(jj)
sys.getsizeof(jj)
sys.getsizeof(j3)

sys.path
sys.maxsize

# time
import time

## 1970년 1월 1일 0시 0분 0초 기준
t1 = time.time()
print(" {} 초".format(t1))
print(" {} 초".format(int(t1)))

## 구조체 : method가 없고 field만 있는 class
t2 = time.gmtime() ## global time
print(t2)
print("{}/{}/{}".format(t2.tm_year, t2.tm_mon, t2.tm_mday))
print("{}:{}:{}".format(t2.tm_hour, t2.tm_min, t2.tm_sec))

t3 = time.localtime() ## 내 컴퓨터 시간
print(t3)
print("{}/{}/{}".format(t3.tm_year, t3.tm_mon, t3.tm_mday))
print("{}:{}:{}".format(t3.tm_hour, t3.tm_min, t3.tm_sec))

start = time.time()
for i in range(1000):
    print("싫어", end = '')
end = time.time()

print("\n")
print( " 소요 시간 : ", end - start)

# datetime
import datetime as dt

now = dt.datetime.now()
print("{}/{}/{}".format(now.year, now.month, now.day))
print("{}:{}:{}".format(now.hour, now.minute, now.second))

import calendar as cld

c = cld.calendar(2018)
print(c)

print(cld.month(2019,5))

def whatwdy():
    week = ["월", "화", "수", "목", "금", "토", "일"]
    print(" 생년월일을 입력하세요.")
    year = int(input(" 연도 > "))
    month = int(input(" 월 > "))
    day = int(input(" 일 > "))
    
    q = cld.weekday(year, month, day)
    print(" 나는 {}요일에 태어났어요.".format(week[q]))

whatwdy()


# file 다루기

f = open("c:\\dd\\mon.txt", "r")
a = f.read()
print(a)
f.close()

f = open("c:\\dd\\33.txt", "w")
f.write("배가 고파요\n")
f.write("배꼽시계가 울려요\n")
f.close()
## 그냥 w로 불러와 쓰면 덮어쓴다. a 사용
f = open("c:\\dd\\33.txt", "a")
f.write("빨리 끝내주세요\n")
f.close()
f = open("c:\\dd\\44.txt", "a")
for i in range(50):
    f.write("배고파")
f.close()

f.write("\n")
for i in range(5):
    name = input(" 이름 > ")
    f.write(name)
    f.write("\n")
    
f.close()

f = open("c:\\dd\\ace.txt", "r")
print(f.read())

# sql 연결

import sqlite3 as sql

conn = sql.connect('c:\\dd\\adb') # db 연결
c = conn.cursor() # 커서 연결

c.execute('drop table if exists Lunch')
c.execute("create table Lunch(menu char(20), price int)")
c.execute("insert into Lunch values('비빔밥', 7000)")
c.execute("insert into Lunch values('볶음밥', 8000)")
c.execute("insert into Lunch values('짜장면', 6000)")

conn.commit()

c.execute("select * from Lunch")
k = c.fetchall()

for i in k:
    print(" {} {}원".format(i[0], i[1]))
    
c.execute('drop table if exists ltx')

c.close() # 커서 닫고
conn.close() # db 닫고

menu = input(" 좋아하는 메뉴 : ")
price = int(input(" 가격 : "))
c.execute("insert into Lunch(menu, price) values(?, ?)", (menu, price))

#################################
a = []

with open("c:\\dd\\ss.txt", 'r') as f:
	for i in range(10):
		a.append(f.readline().split())

for i in range(len(a)):
	for j in range(1,4):
		a[i][j] = int(a[i][j])
	a[i].append(sum(a[i][1:4]))
	a[i].append(round((a[i][4] / 3), 2))

korT = 0
engT = 0
matT = 0

for i in range(len(a)):
	korT += a[i][1]
	engT += a[i][2]
	matT += a[i][3]

korA = korT / len(a)
engA = engT / len(a)
matA = matT / len(a)

everyT = korT + engT + matT
everyA = everyT / (3 * len(a))
#########################

## 성적 처리

conn = sql.connect('c:\\dd\\adb') 
c = conn.cursor()

c.execute("drop table if exists students")
c.execute("create table students(\
        name char(20), kor int, eng int, mat int, tot int, avg int)")

sql1 = "insert into students(name, kor, eng, mat, tot, avg) values(?, ?, ?, ?, ?, ?)"

for i in range(10):
    c.execute(sql1, (a[i][0], a[i][1], a[i][2], a[i][3], a[i][4], round(a[i][5])))

c.execute("select * from students")
conn.commit()
kk = c.fetchall()

f = open('c:\\dd\\source.txt', 'w')

f.write(" *************** 성 적 표 ************** \n")
f.write(" *************************************** \n")
f.write("  이 름  국어  영어  수학  총점    평균  \n")
f.write(" *************************************** \n")

for i in range(len(a)):
	f.write(" {} {:5d} {:5d} {:5d} {:5d} {:7.2f} \n"
              .format(kk[i][0], kk[i][1], kk[i][2], kk[i][3], kk[i][4], kk[i][5]))

f.write(" ************************************** \n")
f.write(" 총평균 {:5.1f} {:5.1f} {:5.1f} {:5d} {:7.2f} \n"
      .format(korA, engA, matA, everyT, everyA))


f.close()

#############################성적 처리 source.txt로 저장
########################################################

# 엑셀 파일 부르기

import openpyxl

exf = openpyxl.load_workbook('c:\\dd\\ss.xlsx')
aws = exf.active

print(" *************** 성 적 표 ************** ")
print(" *************************************** ")
print("  이 름  국어  영어  수학  총점    평균  ")
print(" *************************************** ")

for i in aws.rows:
    index = i[0].row
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    mat = i[3].value
    total = kor + eng + mat
    avg = total / 3
    
    aws.cell(row = index, column = 5).value = total
    aws.cell(row = index, column = 6).value = avg
    
    print(" {} {:5d} {:5d} {:5d} {:5d} {:7.2f} "
              .format(name, kor, eng, mat, total, avg))
    
print(" *************************************** ")
exf.save("c:\\dd\\report.xlsx")
exf.close()


# 예외 처리
g = [j*5 for j in range(10)]
gg = iter(g)
while True:
    try:
        num = next(gg)
    
    except:
        print(" 범위 끝났어 ")
        break
    print(num)
    
a = [1, 2, 3]

try:
    #a[0] / 0
    #a[5] = 6
    a[2] + "good"

except ZeroDivisionError:
    print(" 0으로 못나눠용 ")

except IndexError:
    print(" 인덱스를 벗어났어 ")
    
except TypeError:
    print(" type이 맞지 않아 ")

finally:
    print(" 무조건 실행 ")

# 난수 생성
    
import random

print(" * 0 ~ 1 사이의 값 * ")
print(" ****************** ")
for i in range(10):
    print(random.random())
    
print(" * 1 ~ 1000 사이의 값 * ")
print(" ********************* ")
for i in range(10):
    print(random.randint(1,1000))

## 로또 번호 추출
for i in range(5):
    num = random.sample(range(1,46), 6)
    num.sort()
    print(num)

# os modules
    
import os
os.system("notepad")
os.system("mspaint")
os.sys
os.path
os.mkdir("dir")
os.mkdir("c:\\dd\\OK")
os.mkdir("c:\\dd\\파이썬")

## os 모듈 이용
def osmodules():
    import os
    import random as rd
    
    while True:
        print(" ***** 메 뉴 ***** ")
        print(" ***************** ")
        print("  1. 계산기        ")
        print("  2. 메모장        ")
        print("  3. 명령 프롬프트 ")
        print("  4. 숫자 맞추기   ")
        print("  0. 프로그램 종료 ")
        print(" ***************** ")
        menu = int(input(" 메뉴를 고르세요 > "))
        
        if menu == 0:
            print(" 종료헙니다. ")
            break
            
        if menu == 1:
            print(" 계산기 실행 ")
            os.system('calc')
                
        if menu == 2:
            print(" 메모장 실행 ")
            os.system('calc')
                    
        if menu == 3:
            print(" 명령 프롬프트 실행 ")
            os.system('cmd')

        if menu == 4:
            print(" 숫자 맞추기 실행 ")
            num = int(input(" 1부터 원하는 숫자 범위를 고르세요 > "))
            random_number = rd.randint(1,num)
            
            cnt = 0
            print("1 ~ 100 중에 고르세요!")

            while True:
                pick = int(input("Pick > "))
                cnt += 1
                if pick == random_number:
                    print("{}번 만에 맞추셨습니다!".format(cnt))
                    break

                elif pick > random_number:
                    print("{} 보다는 작습니다!".format(pick))
                        
                else:
                    print("{} 보다는 큰 것 같습니다!".format(pick))
                    continue
                    
        else:
            print(" 존재하지 않는 메뉴입니다. ")
            print(" 다시 고르세요. ")
            continue